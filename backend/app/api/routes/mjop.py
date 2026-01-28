"""MJOP (Maintenance Plan) API routes.

Implements EPIC-014 (MJOP & Onderhoudsplanning):
- FEAT-029: MJOP Import & Beheer (STORY-062, STORY-063, STORY-064)
- FEAT-030: Reserveberekening & Prognose (STORY-065, STORY-066)
- FEAT-031: Onderhoudstaak Beheer (STORY-067, STORY-068)
"""

import json
import uuid
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies.auth import (
    CurrentUser,
    require_beheerder,
    require_member,
)
from app.db.models.models import (
    MaintenanceElement,
    MaintenanceElementCategory,
    MaintenancePriority,
    MaintenanceStatus,
    MaintenanceTask,
    MJOPImportBatch,
    VVE,
)
from app.db.session import get_db
from app.schemas.mjop import (
    MaintenanceElementCreate,
    MaintenanceElementResponse,
    MaintenanceElementUpdate,
    MaintenanceTaskCreate,
    MaintenanceTaskResponse,
    MaintenanceTaskUpdate,
    MJOPImportBatchResponse,
    MJOPImportPreviewResponse,
    MJOPImportPreviewRow,
    MJOPImportRequest,
    MJOPImportResponse,
    MJOPTimelineResponse,
    ReserveCalculationRequest,
    ReserveCalculationResponse,
    TimelineItem,
    WhatIfScenarioRequest,
    WhatIfScenarioResponse,
    WhatIfYearProjection,
)

router = APIRouter(prefix="/vves/{vve_id}/mjop", tags=["mjop"])

# Field mapping for Excel import
IMPORTABLE_FIELDS = {
    "name": {"required": True, "type": "string"},
    "description": {"required": False, "type": "string"},
    "category": {"required": True, "type": "category"},
    "location": {"required": False, "type": "string"},
    "quantity": {"required": False, "type": "int"},
    "unit": {"required": False, "type": "string"},
    "installation_year": {"required": False, "type": "int"},
    "expected_lifespan_years": {"required": False, "type": "int"},
    "last_maintenance_year": {"required": False, "type": "int"},
    "next_maintenance_year": {"required": False, "type": "int"},
    "estimated_cost": {"required": False, "type": "decimal"},
    "priority": {"required": False, "type": "priority"},
}

# Dutch to English category mapping
CATEGORY_MAPPING = {
    "dak": MaintenanceElementCategory.ROOF,
    "roof": MaintenanceElementCategory.ROOF,
    "gevel": MaintenanceElementCategory.FACADE,
    "facade": MaintenanceElementCategory.FACADE,
    "fundering": MaintenanceElementCategory.FOUNDATION,
    "foundation": MaintenanceElementCategory.FOUNDATION,
    "ramen": MaintenanceElementCategory.WINDOWS,
    "windows": MaintenanceElementCategory.WINDOWS,
    "deuren": MaintenanceElementCategory.DOORS,
    "doors": MaintenanceElementCategory.DOORS,
    "lift": MaintenanceElementCategory.ELEVATOR,
    "elevator": MaintenanceElementCategory.ELEVATOR,
    "verwarming": MaintenanceElementCategory.HEATING,
    "heating": MaintenanceElementCategory.HEATING,
    "leidingwerk": MaintenanceElementCategory.PLUMBING,
    "plumbing": MaintenanceElementCategory.PLUMBING,
    "elektra": MaintenanceElementCategory.ELECTRICAL,
    "electrical": MaintenanceElementCategory.ELECTRICAL,
    "gemeenschappelijk": MaintenanceElementCategory.COMMON_AREAS,
    "common_areas": MaintenanceElementCategory.COMMON_AREAS,
    "tuin": MaintenanceElementCategory.GARDEN,
    "garden": MaintenanceElementCategory.GARDEN,
    "parkeren": MaintenanceElementCategory.PARKING,
    "parking": MaintenanceElementCategory.PARKING,
    "overig": MaintenanceElementCategory.OTHER,
    "other": MaintenanceElementCategory.OTHER,
}

PRIORITY_MAPPING = {
    "laag": MaintenancePriority.LOW,
    "low": MaintenancePriority.LOW,
    "gemiddeld": MaintenancePriority.MEDIUM,
    "medium": MaintenancePriority.MEDIUM,
    "hoog": MaintenancePriority.HIGH,
    "high": MaintenancePriority.HIGH,
    "urgent": MaintenancePriority.URGENT,
    "spoed": MaintenancePriority.URGENT,
}


def parse_cell_value(value: Any, field_type: str) -> tuple[Any, str | None]:
    """Parse a cell value based on expected type. Returns (value, error)."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None, None

    try:
        if field_type == "string":
            return str(value).strip(), None
        elif field_type == "int":
            return int(float(value)), None
        elif field_type == "decimal":
            return Decimal(str(value)), None
        elif field_type == "category":
            val = str(value).strip().lower()
            if val in CATEGORY_MAPPING:
                return CATEGORY_MAPPING[val], None
            return None, f"Onbekende categorie: {value}"
        elif field_type == "priority":
            val = str(value).strip().lower()
            if val in PRIORITY_MAPPING:
                return PRIORITY_MAPPING[val], None
            return MaintenancePriority.MEDIUM, None  # Default
    except (ValueError, TypeError) as e:
        return None, f"Ongeldige waarde: {value} ({str(e)})"

    return None, f"Onbekend type: {field_type}"


def suggest_column_mapping(headers: list[str]) -> dict[str, str]:
    """Suggest column mapping based on header names."""
    suggestions: dict[str, str] = {}
    header_lower = [h.lower() if h else "" for h in headers]

    mapping_hints = {
        "name": ["naam", "name", "element", "onderdeel", "omschrijving"],
        "description": ["beschrijving", "description", "toelichting", "details"],
        "category": ["categorie", "category", "type", "soort"],
        "location": ["locatie", "location", "plaats", "positie"],
        "quantity": ["aantal", "quantity", "hoeveelheid", "stuks"],
        "unit": ["eenheid", "unit", "meeteenheid"],
        "installation_year": ["bouwjaar", "installation_year", "jaar van installatie", "aanlegjaar"],
        "expected_lifespan_years": ["levensduur", "lifespan", "verwachte levensduur"],
        "last_maintenance_year": ["laatste onderhoud", "last_maintenance", "laatst onderhouden"],
        "next_maintenance_year": ["volgende onderhoud", "next_maintenance", "gepland", "planning"],
        "estimated_cost": ["kosten", "cost", "prijs", "bedrag", "geschatte kosten"],
        "priority": ["prioriteit", "priority", "urgentie"],
    }

    for field, hints in mapping_hints.items():
        for i, header in enumerate(header_lower):
            if any(hint in header for hint in hints):
                # Use Excel column letter
                col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
                suggestions[field] = col_letter
                break

    return suggestions


# ============================================================================
# Excel Import Endpoints (STORY-062)
# ============================================================================


@router.post(
    "/import/upload",
    response_model=MJOPImportPreviewResponse,
    summary="Upload Excel voor MJOP preview",
    description="""
    STORY-062: Upload een Excel bestand (.xlsx) met MJOP data.
    
    Retourneert een preview van de data met voorgestelde kolom-mapping.
    De kolommen worden automatisch gedetecteerd en gemapt waar mogelijk.
    """,
)
async def upload_mjop_excel(
    vve_id: uuid.UUID,
    file: Annotated[UploadFile, File(description="Excel bestand (.xlsx)")],
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MJOPImportPreviewResponse:
    """Upload Excel file and return preview with column detection."""
    # Verify VVE exists
    vve_result = await db.execute(select(VVE).where(VVE.id == vve_id))
    if vve_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VVE niet gevonden",
        )

    # Validate file type
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Alleen .xlsx bestanden zijn toegestaan",
        )

    try:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Geen actief werkblad gevonden in Excel bestand",
            )

        # Get headers from first row
        headers: list[str] = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")

        # Suggest column mapping
        suggested_mapping = suggest_column_mapping(headers)

        # Preview first 10 rows
        preview_rows: list[MJOPImportPreviewRow] = []
        total_rows = 0
        valid_rows = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue  # Skip empty rows

            total_rows += 1
            row_data = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
            errors: list[str] = []

            # Validate required fields based on suggested mapping
            for field, config in IMPORTABLE_FIELDS.items():
                if config["required"] and field in suggested_mapping:
                    col_letter = suggested_mapping[field]
                    col_idx = ord(col_letter) - 65
                    if col_idx < len(row) and (row[col_idx] is None or str(row[col_idx]).strip() == ""):
                        errors.append(f"Verplicht veld '{field}' ontbreekt")

            is_valid = len(errors) == 0
            if is_valid:
                valid_rows += 1

            if len(preview_rows) < 10:  # Only preview first 10 rows
                preview_rows.append(
                    MJOPImportPreviewRow(
                        row_number=row_idx,
                        data=row_data,
                        errors=errors,
                        is_valid=is_valid,
                    )
                )

        wb.close()

        return MJOPImportPreviewResponse(
            filename=file.filename or "unknown.xlsx",
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=total_rows - valid_rows,
            preview_rows=preview_rows,
            detected_columns=headers,
            suggested_mapping=suggested_mapping,
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Fout bij verwerken Excel bestand: {str(e)}",
        )


@router.post(
    "/import/confirm",
    response_model=MJOPImportResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Bevestig en voer MJOP import uit",
    description="""
    STORY-062: Bevestig de kolom-mapping en importeer de MJOP data.
    
    Na preview kan de gebruiker de mapping aanpassen en import starten.
    """,
)
async def confirm_mjop_import(
    vve_id: uuid.UUID,
    file: Annotated[UploadFile, File(description="Excel bestand (.xlsx)")],
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
    column_mapping: str = Query(..., description="JSON string met kolom-mapping"),
    skip_invalid_rows: bool = Query(False, description="Sla ongeldige rijen over"),
) -> MJOPImportResponse:
    """Execute MJOP import with confirmed column mapping."""
    # Verify VVE exists
    vve_result = await db.execute(select(VVE).where(VVE.id == vve_id))
    if vve_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VVE niet gevonden",
        )

    try:
        mapping = json.loads(column_mapping)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ongeldige kolom-mapping JSON",
        )

    # Validate file
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Alleen .xlsx bestanden zijn toegestaan",
        )

    try:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Geen actief werkblad gevonden",
            )

        # Create import batch
        batch = MJOPImportBatch(
            vve_id=vve_id,
            filename=file.filename or "unknown.xlsx",
            column_mapping=json.dumps(mapping),
            created_by_id=current_user.id,
        )
        db.add(batch)
        await db.flush()

        # Process rows
        total_rows = 0
        imported_rows = 0
        failed_rows = 0
        errors: list[dict[str, Any]] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            total_rows += 1
            row_errors: list[str] = []
            element_data: dict[str, Any] = {}

            # Extract and parse values based on mapping
            for field, col_letter in mapping.items():
                if field not in IMPORTABLE_FIELDS:
                    continue

                col_idx = ord(col_letter.upper()) - 65
                if col_idx >= len(row):
                    continue

                cell_value = row[col_idx]
                field_type = IMPORTABLE_FIELDS[field]["type"]
                parsed_value, error = parse_cell_value(cell_value, field_type)

                if error:
                    row_errors.append(f"{field}: {error}")
                elif parsed_value is not None:
                    element_data[field] = parsed_value

            # Check required fields
            for field, config in IMPORTABLE_FIELDS.items():
                if config["required"] and field not in element_data:
                    row_errors.append(f"Verplicht veld '{field}' ontbreekt of is ongeldig")

            if row_errors:
                failed_rows += 1
                errors.append({"row": row_idx, "errors": row_errors})
                if not skip_invalid_rows:
                    continue
            else:
                # Create maintenance element
                element = MaintenanceElement(
                    vve_id=vve_id,
                    name=element_data.get("name", ""),
                    description=element_data.get("description"),
                    category=element_data.get("category", MaintenanceElementCategory.OTHER),
                    location=element_data.get("location"),
                    quantity=element_data.get("quantity", 1),
                    unit=element_data.get("unit"),
                    installation_year=element_data.get("installation_year"),
                    expected_lifespan_years=element_data.get("expected_lifespan_years"),
                    last_maintenance_year=element_data.get("last_maintenance_year"),
                    next_maintenance_year=element_data.get("next_maintenance_year"),
                    estimated_cost=element_data.get("estimated_cost"),
                    priority=element_data.get("priority", MaintenancePriority.MEDIUM),
                    import_batch_id=batch.id,
                    import_row_number=row_idx,
                    created_by_id=current_user.id,
                )
                db.add(element)
                imported_rows += 1

        # Update batch stats
        batch.total_rows = total_rows
        batch.imported_rows = imported_rows
        batch.failed_rows = failed_rows
        batch.is_completed = True
        batch.error_log = json.dumps(errors) if errors else None

        await db.commit()
        wb.close()

        return MJOPImportResponse(
            batch_id=batch.id,
            filename=file.filename or "unknown.xlsx",
            total_rows=total_rows,
            imported_rows=imported_rows,
            failed_rows=failed_rows,
            errors=errors[:20],  # Return first 20 errors
        )

    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Fout bij importeren: {str(e)}",
        )


@router.get(
    "/import/batches",
    response_model=list[MJOPImportBatchResponse],
    summary="Lijst van import batches",
)
async def list_import_batches(
    vve_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
) -> list[MJOPImportBatchResponse]:
    """Get list of MJOP import batches for audit."""
    query = (
        select(MJOPImportBatch)
        .where(MJOPImportBatch.vve_id == vve_id)
        .order_by(MJOPImportBatch.created_at.desc())
        .offset(skip)
        .limit(limit)
    )
    result = await db.execute(query)
    batches = result.scalars().all()
    return [MJOPImportBatchResponse.model_validate(b) for b in batches]


# ============================================================================
# Maintenance Element CRUD (STORY-063)
# ============================================================================


@router.post(
    "/elements",
    response_model=MaintenanceElementResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Onderhoudselement handmatig toevoegen",
    description="""
    STORY-063: Als beheerder wil ik handmatig onderhoudselementen kunnen toevoegen.
    """,
)
async def create_maintenance_element(
    vve_id: uuid.UUID,
    element_data: MaintenanceElementCreate,
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MaintenanceElementResponse:
    """Create a new maintenance element."""
    # Verify VVE exists
    vve_result = await db.execute(select(VVE).where(VVE.id == vve_id))
    if vve_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VVE niet gevonden",
        )

    element = MaintenanceElement(
        vve_id=vve_id,
        created_by_id=current_user.id,
        **element_data.model_dump(),
    )
    db.add(element)
    await db.commit()
    await db.refresh(element)

    return MaintenanceElementResponse.model_validate(element)


@router.get(
    "/elements",
    response_model=list[MaintenanceElementResponse],
    summary="Lijst onderhoudselementen",
)
async def list_maintenance_elements(
    vve_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    category: MaintenanceElementCategory | None = None,
) -> list[MaintenanceElementResponse]:
    """Get list of maintenance elements for a VVE."""
    query = select(MaintenanceElement).where(MaintenanceElement.vve_id == vve_id)

    if category:
        query = query.where(MaintenanceElement.category == category)

    query = query.order_by(MaintenanceElement.next_maintenance_year.asc().nulls_last())
    query = query.offset(skip).limit(limit)

    result = await db.execute(query)
    elements = result.scalars().all()
    return [MaintenanceElementResponse.model_validate(e) for e in elements]


@router.get(
    "/elements/{element_id}",
    response_model=MaintenanceElementResponse,
    summary="Onderhoudselement details",
)
async def get_maintenance_element(
    vve_id: uuid.UUID,
    element_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MaintenanceElementResponse:
    """Get a specific maintenance element."""
    result = await db.execute(
        select(MaintenanceElement).where(
            MaintenanceElement.id == element_id,
            MaintenanceElement.vve_id == vve_id,
        )
    )
    element = result.scalar_one_or_none()

    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudselement niet gevonden",
        )

    return MaintenanceElementResponse.model_validate(element)


@router.put(
    "/elements/{element_id}",
    response_model=MaintenanceElementResponse,
    summary="Onderhoudselement wijzigen",
)
async def update_maintenance_element(
    vve_id: uuid.UUID,
    element_id: uuid.UUID,
    update_data: MaintenanceElementUpdate,
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MaintenanceElementResponse:
    """Update a maintenance element."""
    result = await db.execute(
        select(MaintenanceElement).where(
            MaintenanceElement.id == element_id,
            MaintenanceElement.vve_id == vve_id,
        )
    )
    element = result.scalar_one_or_none()

    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudselement niet gevonden",
        )

    update_dict = update_data.model_dump(exclude_unset=True)
    for field, value in update_dict.items():
        setattr(element, field, value)

    await db.commit()
    await db.refresh(element)

    return MaintenanceElementResponse.model_validate(element)


@router.delete(
    "/elements/{element_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Onderhoudselement verwijderen",
)
async def delete_maintenance_element(
    vve_id: uuid.UUID,
    element_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    """Delete a maintenance element."""
    result = await db.execute(
        select(MaintenanceElement).where(
            MaintenanceElement.id == element_id,
            MaintenanceElement.vve_id == vve_id,
        )
    )
    element = result.scalar_one_or_none()

    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudselement niet gevonden",
        )

    await db.delete(element)
    await db.commit()


# ============================================================================
# Maintenance Task CRUD (STORY-067, STORY-068)
# ============================================================================


@router.post(
    "/tasks",
    response_model=MaintenanceTaskResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Onderhoudstaak aanmaken",
    description="""
    STORY-067: Als beheerder wil ik een onderhoudstaak kunnen aanmaken en toewijzen.
    """,
)
async def create_maintenance_task(
    vve_id: uuid.UUID,
    task_data: MaintenanceTaskCreate,
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MaintenanceTaskResponse:
    """Create a new maintenance task."""
    # Verify element exists and belongs to VVE
    element_result = await db.execute(
        select(MaintenanceElement).where(
            MaintenanceElement.id == task_data.element_id,
            MaintenanceElement.vve_id == vve_id,
        )
    )
    if element_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudselement niet gevonden",
        )

    task = MaintenanceTask(
        vve_id=vve_id,
        created_by_id=current_user.id,
        **task_data.model_dump(),
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)

    return MaintenanceTaskResponse.model_validate(task)


@router.get(
    "/tasks",
    response_model=list[MaintenanceTaskResponse],
    summary="Lijst onderhoudstaken",
)
async def list_maintenance_tasks(
    vve_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    status_filter: MaintenanceStatus | None = Query(None, alias="status"),
    element_id: uuid.UUID | None = None,
) -> list[MaintenanceTaskResponse]:
    """Get list of maintenance tasks."""
    query = select(MaintenanceTask).where(MaintenanceTask.vve_id == vve_id)

    if status_filter:
        query = query.where(MaintenanceTask.status == status_filter)
    if element_id:
        query = query.where(MaintenanceTask.element_id == element_id)

    query = query.order_by(MaintenanceTask.planned_year.asc().nulls_last())
    query = query.offset(skip).limit(limit)

    result = await db.execute(query)
    tasks = result.scalars().all()
    return [MaintenanceTaskResponse.model_validate(t) for t in tasks]


@router.get(
    "/tasks/{task_id}",
    response_model=MaintenanceTaskResponse,
    summary="Onderhoudstaak details",
)
async def get_maintenance_task(
    vve_id: uuid.UUID,
    task_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MaintenanceTaskResponse:
    """Get a specific maintenance task."""
    result = await db.execute(
        select(MaintenanceTask).where(
            MaintenanceTask.id == task_id,
            MaintenanceTask.vve_id == vve_id,
        )
    )
    task = result.scalar_one_or_none()

    if task is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudstaak niet gevonden",
        )

    return MaintenanceTaskResponse.model_validate(task)


@router.put(
    "/tasks/{task_id}",
    response_model=MaintenanceTaskResponse,
    summary="Onderhoudstaak status bijwerken",
    description="""
    STORY-068: Als beheerder wil ik de status van een onderhoudstaak kunnen bijwerken.
    """,
)
async def update_maintenance_task(
    vve_id: uuid.UUID,
    task_id: uuid.UUID,
    update_data: MaintenanceTaskUpdate,
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> MaintenanceTaskResponse:
    """Update a maintenance task."""
    result = await db.execute(
        select(MaintenanceTask).where(
            MaintenanceTask.id == task_id,
            MaintenanceTask.vve_id == vve_id,
        )
    )
    task = result.scalar_one_or_none()

    if task is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudstaak niet gevonden",
        )

    update_dict = update_data.model_dump(exclude_unset=True)
    for field, value in update_dict.items():
        setattr(task, field, value)

    await db.commit()
    await db.refresh(task)

    return MaintenanceTaskResponse.model_validate(task)


@router.delete(
    "/tasks/{task_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Onderhoudstaak verwijderen",
)
async def delete_maintenance_task(
    vve_id: uuid.UUID,
    task_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_beheerder)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> None:
    """Delete a maintenance task."""
    result = await db.execute(
        select(MaintenanceTask).where(
            MaintenanceTask.id == task_id,
            MaintenanceTask.vve_id == vve_id,
        )
    )
    task = result.scalar_one_or_none()

    if task is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Onderhoudstaak niet gevonden",
        )

    await db.delete(task)
    await db.commit()


# ============================================================================
# Timeline Visualization (STORY-064)
# ============================================================================


@router.get(
    "/timeline",
    response_model=MJOPTimelineResponse,
    summary="MJOP timeline visualisatie",
    description="""
    STORY-064: Visualisatie van gepland onderhoud op tijdlijn.
    """,
)
async def get_mjop_timeline(
    vve_id: uuid.UUID,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
    start_year: int = Query(2024, ge=2000, le=2100),
    end_year: int = Query(2034, ge=2000, le=2100),
) -> MJOPTimelineResponse:
    """Get MJOP timeline data for visualization."""
    # Verify VVE exists
    vve_result = await db.execute(select(VVE).where(VVE.id == vve_id))
    if vve_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VVE niet gevonden",
        )

    # Get elements with maintenance planned in range
    query = select(MaintenanceElement).where(
        MaintenanceElement.vve_id == vve_id,
        MaintenanceElement.next_maintenance_year >= start_year,
        MaintenanceElement.next_maintenance_year <= end_year,
    )
    result = await db.execute(query)
    elements = result.scalars().all()

    # Get tasks for these elements
    element_ids = [e.id for e in elements]
    tasks_query = select(MaintenanceTask).where(MaintenanceTask.element_id.in_(element_ids))
    tasks_result = await db.execute(tasks_query)
    tasks = {t.element_id: t for t in tasks_result.scalars().all()}

    # Build timeline items
    items: list[TimelineItem] = []
    total_by_year: dict[int, Decimal] = {}
    total_by_category: dict[str, Decimal] = {}

    for element in elements:
        year = element.next_maintenance_year
        if year is None:
            continue

        cost = element.estimated_cost or Decimal("0")
        task = tasks.get(element.id)

        items.append(
            TimelineItem(
                element_id=element.id,
                element_name=element.name,
                category=element.category,
                year=year,
                estimated_cost=cost,
                priority=element.priority,
                has_task=task is not None,
                task_status=task.status if task else None,
            )
        )

        # Aggregate by year
        if year not in total_by_year:
            total_by_year[year] = Decimal("0")
        total_by_year[year] += cost

        # Aggregate by category
        cat_key = element.category.value
        if cat_key not in total_by_category:
            total_by_category[cat_key] = Decimal("0")
        total_by_category[cat_key] += cost

    return MJOPTimelineResponse(
        vve_id=vve_id,
        start_year=start_year,
        end_year=end_year,
        items=items,
        total_by_year=total_by_year,
        total_by_category=total_by_category,
    )


# ============================================================================
# Reserve Calculation (STORY-065, STORY-066)
# ============================================================================


@router.post(
    "/reserve-calculation",
    response_model=ReserveCalculationResponse,
    summary="Reserveberekening automatisch",
    description="""
    STORY-065: Automatische berekening van benodigde reserves op basis van MJOP.
    """,
)
async def calculate_reserves(
    vve_id: uuid.UUID,
    request: ReserveCalculationRequest,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> ReserveCalculationResponse:
    """Calculate required reserves based on MJOP elements."""
    # Verify VVE exists
    vve_result = await db.execute(select(VVE).where(VVE.id == vve_id))
    if vve_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VVE niet gevonden",
        )

    current_year = datetime.now().year
    end_year = current_year + request.years_ahead

    # Get elements with maintenance in range
    query = select(MaintenanceElement).where(
        MaintenanceElement.vve_id == vve_id,
        MaintenanceElement.next_maintenance_year >= current_year,
        MaintenanceElement.next_maintenance_year <= end_year,
    )
    result = await db.execute(query)
    elements = result.scalars().all()

    by_year: dict[int, Decimal] = {}
    by_category: dict[str, Decimal] = {}
    total_required = Decimal("0")

    for element in elements:
        year = element.next_maintenance_year
        cost = element.estimated_cost or Decimal("0")
        
        if year is None:
            continue

        total_required += cost

        if year not in by_year:
            by_year[year] = Decimal("0")
        by_year[year] += cost

        cat_key = element.category.value
        if cat_key not in by_category:
            by_category[cat_key] = Decimal("0")
        by_category[cat_key] += cost

    # Add contingency
    contingency_amount = Decimal("0")
    if request.include_contingency:
        contingency_amount = total_required * (request.contingency_percentage / 100)
        total_required += contingency_amount

    # Calculate annual contribution
    annual_contribution = total_required / request.years_ahead if request.years_ahead > 0 else Decimal("0")

    return ReserveCalculationResponse(
        vve_id=vve_id,
        years_ahead=request.years_ahead,
        total_required=total_required,
        annual_contribution=annual_contribution,
        by_year=by_year,
        by_category=by_category,
        contingency_amount=contingency_amount,
    )


# ============================================================================
# What-If Scenario Calculation (STORY-066)
# ============================================================================


@router.post(
    "/what-if-scenario",
    response_model=WhatIfScenarioResponse,
    summary="What-if scenario doorrekenen",
    description="""
    STORY-066: Als penningmeester wil ik what-if scenario's kunnen doorrekenen
    met verschillende contributiehoogtes, zodat ik de impact op reserves kan
    presenteren aan de ALV.
    
    Features:
    - Adjust contribution per owner (percentage slider)
    - Compare current vs scenario projections
    - Year-by-year breakdown for graphing
    - Save scenario for presentation
    - Export to PDF ready data
    """,
)
async def calculate_what_if_scenario(
    vve_id: uuid.UUID,
    request: WhatIfScenarioRequest,
    current_user: Annotated[CurrentUser, Depends(require_member)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> WhatIfScenarioResponse:
    """Calculate what-if scenario comparing current vs adjusted projections."""
    # Verify VVE exists
    vve_result = await db.execute(select(VVE).where(VVE.id == vve_id))
    if vve_result.scalar_one_or_none() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VVE niet gevonden",
        )

    current_year = datetime.now().year
    end_year = current_year + request.years_ahead

    # Get elements with maintenance in range
    query = select(MaintenanceElement).where(
        MaintenanceElement.vve_id == vve_id,
        MaintenanceElement.next_maintenance_year >= current_year,
        MaintenanceElement.next_maintenance_year <= end_year,
    )
    result = await db.execute(query)
    elements = list(result.scalars().all())

    # Calculate original projections
    original_by_year: dict[int, Decimal] = {}
    original_by_category: dict[str, Decimal] = {}
    original_total = Decimal("0")

    for element in elements:
        year = element.next_maintenance_year
        cost = element.estimated_cost or Decimal("0")
        
        if year is None:
            continue

        original_total += cost
        if year not in original_by_year:
            original_by_year[year] = Decimal("0")
        original_by_year[year] += cost

        cat_key = element.category.value
        if cat_key not in original_by_category:
            original_by_category[cat_key] = Decimal("0")
        original_by_category[cat_key] += cost

    # Calculate scenario projections with adjustments
    scenario_by_year: dict[int, Decimal] = {}
    scenario_by_category: dict[str, Decimal] = {}
    scenario_total = Decimal("0")
    postponed_elements = set(request.postpone_elements)

    for element in elements:
        year = element.next_maintenance_year
        cost = element.estimated_cost or Decimal("0")
        
        if year is None:
            continue

        # Apply cost increase percentage
        adjusted_cost = cost * (1 + request.cost_increase_percentage / 100)

        # Apply postponement if element is in postpone list
        adjusted_year = year
        if element.id in postponed_elements:
            adjusted_year = year + request.postpone_years

        # Only count if still within projection range
        if adjusted_year <= end_year:
            scenario_total += adjusted_cost
            if adjusted_year not in scenario_by_year:
                scenario_by_year[adjusted_year] = Decimal("0")
            scenario_by_year[adjusted_year] += adjusted_cost

            cat_key = element.category.value
            if cat_key not in scenario_by_category:
                scenario_by_category[cat_key] = Decimal("0")
            scenario_by_category[cat_key] += adjusted_cost

    # Add contingency if requested
    if request.include_contingency:
        original_contingency = original_total * (request.contingency_percentage / 100)
        scenario_contingency = scenario_total * (request.contingency_percentage / 100)
        original_total += original_contingency
        scenario_total += scenario_contingency

    # Calculate annual contributions
    annual_contribution_original = (
        original_total / request.years_ahead if request.years_ahead > 0 else Decimal("0")
    )
    
    # Apply contribution adjustment to scenario
    contribution_multiplier = 1 + (request.contribution_adjustment_percentage / 100)
    annual_contribution_scenario = annual_contribution_original * contribution_multiplier

    # Calculate difference
    difference = scenario_total - original_total
    difference_percentage = (
        (difference / original_total * 100) if original_total > 0 else Decimal("0")
    )

    # Generate yearly projections for graphing
    yearly_projections: list[WhatIfYearProjection] = []
    original_balance = Decimal("0")  # Starting balance (could be fetched from VVE)
    scenario_balance = Decimal("0")
    warnings: list[str] = []

    for year in range(current_year, end_year + 1):
        original_cost_this_year = original_by_year.get(year, Decimal("0"))
        scenario_cost_this_year = scenario_by_year.get(year, Decimal("0"))

        # Add contributions, subtract costs
        original_balance = original_balance + annual_contribution_original - original_cost_this_year
        scenario_balance = scenario_balance + annual_contribution_scenario - scenario_cost_this_year

        yearly_projections.append(
            WhatIfYearProjection(
                year=year,
                original_cost=original_cost_this_year,
                scenario_cost=scenario_cost_this_year,
                original_contribution=annual_contribution_original,
                scenario_contribution=annual_contribution_scenario,
                original_reserve_balance=original_balance,
                scenario_reserve_balance=scenario_balance,
            )
        )

        # Check for warnings
        if scenario_balance < 0:
            warnings.append(
                f"Waarschuwing: Negatief saldo in {year} (€{scenario_balance:,.2f})"
            )

    return WhatIfScenarioResponse(
        scenario_name=request.name,
        years_ahead=request.years_ahead,
        original_total=original_total,
        scenario_total=scenario_total,
        difference=difference,
        difference_percentage=difference_percentage,
        annual_contribution_original=annual_contribution_original,
        annual_contribution_scenario=annual_contribution_scenario,
        yearly_projections=yearly_projections,
        by_category_original=original_by_category,
        by_category_scenario=scenario_by_category,
        warnings=warnings,
    )
